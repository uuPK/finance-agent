import csv
import io
import json
from uuid import uuid4

import sqlglot
from openpyxl import load_workbook

from app.services.query_export_service import (
    QueryExportError,
    QueryExportService,
    QueryExportSource,
)
from app.services.sql_executor import SQLExecutionResult


class FakeRepository:
    def __init__(self, source: QueryExportSource | None) -> None:
        self.source = source
        self.finished: list[dict] = []

    def get_source(self, query_id, user_id):
        if self.source and self.source.query_id == query_id and self.source.user_id == user_id:
            return self.source
        return None

    def start_export(self, query_id, user_id, export_format):
        return uuid4()

    def finish_export(self, export_id, **details):
        self.finished.append(details)


class FakeExecutor:
    def __init__(self, result: SQLExecutionResult) -> None:
        self.result = result
        self.sql: str | None = None

    def execute(self, sql: str) -> SQLExecutionResult:
        self.sql = sql
        return self.result


def build_service(export_format: str = "xlsx"):
    query_id = uuid4()
    source = QueryExportSource(
        query_id=query_id,
        user_id="client-1",
        question="查询客户",
        status="completed",
        final_sql="with base as (select 1 as id limit 5) select id from base limit 100",
    )
    result = SQLExecutionResult(
        status="success",
        sql=source.final_sql,
        columns=["id", "note"],
        rows=[{"id": 1, "note": "=SUM(1,1)"}, {"id": 2, "note": "普通文本"}],
        row_count=2,
        truncated=False,
    )
    repository = FakeRepository(source)
    executor = FakeExecutor(result)
    service = QueryExportService(repository=repository, executor=executor)
    artifact = service.export(query_id, "client-1", export_format)
    return artifact, repository, executor


def test_export_removes_only_outer_preview_limit() -> None:
    _, repository, executor = build_service("json")

    expression = sqlglot.parse_one(executor.sql, read="postgres")
    cte_select = next(expression.find_all(sqlglot.exp.CTE)).this

    assert expression.args.get("limit") is None
    assert cte_select.args.get("limit") is not None
    assert repository.finished[0]["status"] == "completed"


def test_xlsx_export_contains_result_and_metadata_sheets() -> None:
    artifact, _, _ = build_service("xlsx")

    workbook = load_workbook(io.BytesIO(artifact.content), read_only=True)

    assert workbook.sheetnames == ["Query Result", "Export Info"]
    assert workbook["Query Result"]["A2"].value == 1
    assert workbook["Query Result"]["B2"].value == "'=SUM(1,1)"
    assert workbook["Export Info"]["B4"].value == 2
    assert artifact.filename.endswith(".xlsx")


def test_csv_and_json_exports_are_business_friendly() -> None:
    csv_artifact, _, _ = build_service("csv")
    rows = list(csv.DictReader(io.StringIO(csv_artifact.content.decode("utf-8-sig"))))
    json_artifact, _, _ = build_service("json")
    payload = json.loads(json_artifact.content)

    assert rows[0]["note"] == "'=SUM(1,1)"
    assert payload["row_count"] == 2
    assert payload["columns"] == ["id", "note"]
    assert len(payload["rows"]) == 2


def test_export_requires_query_ownership() -> None:
    service = QueryExportService(
        repository=FakeRepository(None),
        executor=FakeExecutor(SQLExecutionResult(status="success", sql="select 1")),
    )

    try:
        service.export(uuid4(), "another-client", "xlsx")
    except QueryExportError as exc:
        assert exc.status_code == 404
    else:
        raise AssertionError("Expected query ownership validation to reject the export.")

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from datetime import UTC, datetime
from time import perf_counter
from typing import Any, Literal
from uuid import UUID

import sqlglot
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.engine import Engine

from app.core.config import get_settings
from app.db.session import engine as default_engine
from app.services.sql_executor import SQLExecutionResult, SQLExecutor

ExportFormat = Literal["xlsx", "csv", "json"]


class QueryExportError(Exception):
    def __init__(self, message: str, status_code: int = 400) -> None:
        super().__init__(message)
        self.status_code = status_code


@dataclass(slots=True)
class QueryExportSource:
    query_id: UUID
    user_id: str
    question: str
    status: str
    final_sql: str | None


@dataclass(slots=True)
class QueryExportArtifact:
    content: bytes
    media_type: str
    filename: str
    row_count: int
    truncated: bool


class QueryExportRepository:
    def __init__(self, engine: Engine | None = None) -> None:
        self.engine = engine or default_engine

    def get_source(self, query_id: UUID, user_id: str) -> QueryExportSource | None:
        with self.engine.connect() as connection:
            row = connection.execute(
                text(
                    """
                    select query_id, user_id, question, status, final_sql
                    from agent.query_runs
                    where query_id = :query_id and user_id = :user_id
                    """
                ),
                {"query_id": str(query_id), "user_id": user_id},
            ).mappings().first()
        return QueryExportSource(**dict(row)) if row else None

    def start_export(self, query_id: UUID, user_id: str, export_format: str) -> UUID:
        with self.engine.begin() as connection:
            return connection.execute(
                text(
                    """
                    insert into agent.query_exports (query_id, user_id, export_format)
                    values (:query_id, :user_id, :export_format)
                    returning export_id
                    """
                ),
                {
                    "query_id": str(query_id),
                    "user_id": user_id,
                    "export_format": export_format,
                },
            ).scalar_one()

    def finish_export(
        self,
        export_id: UUID,
        *,
        status: str,
        row_count: int = 0,
        truncated: bool = False,
        elapsed_ms: int = 0,
        error_type: str | None = None,
        error_message: str | None = None,
    ) -> None:
        with self.engine.begin() as connection:
            connection.execute(
                text(
                    """
                    update agent.query_exports
                    set export_status = :status, row_count = :row_count, truncated = :truncated,
                        elapsed_ms = :elapsed_ms, error_type = :error_type,
                        error_message = :error_message, finished_at = now()
                    where export_id = :export_id
                    """
                ),
                {
                    "export_id": str(export_id),
                    "status": status,
                    "row_count": row_count,
                    "truncated": truncated,
                    "elapsed_ms": elapsed_ms,
                    "error_type": error_type,
                    "error_message": error_message,
                },
            )


class QueryExportService:
    media_types = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv; charset=utf-8",
        "json": "application/json; charset=utf-8",
    }

    def __init__(
        self,
        repository: QueryExportRepository | None = None,
        executor: SQLExecutor | None = None,
    ) -> None:
        settings = get_settings()
        self.repository = repository or QueryExportRepository()
        self.max_rows = settings.export_max_rows
        self.executor = executor or SQLExecutor(
            timeout_seconds=settings.export_timeout_seconds,
            max_result_rows=self.max_rows,
        )

    def export(
        self,
        query_id: UUID,
        user_id: str,
        export_format: ExportFormat,
    ) -> QueryExportArtifact:
        source = self.repository.get_source(query_id, user_id)
        if source is None:
            raise QueryExportError("Query run not found for this user.", status_code=404)
        if source.status != "completed" or not source.final_sql:
            raise QueryExportError(
                "Only completed queries with reviewed SQL can be exported.",
                status_code=409,
            )

        export_id = self.repository.start_export(query_id, user_id, export_format)
        started_at = perf_counter()
        try:
            export_sql = self._build_export_sql(source.final_sql)
            result = self.executor.execute(export_sql)
            if result.status != "success":
                raise QueryExportError(
                    result.error_message or "The export query could not be executed.",
                    status_code=422,
                )
            artifact = self._serialize(source, result, export_format)
            self.repository.finish_export(
                export_id,
                status="completed",
                row_count=artifact.row_count,
                truncated=artifact.truncated,
                elapsed_ms=int((perf_counter() - started_at) * 1000),
            )
            return artifact
        except Exception as exc:
            self.repository.finish_export(
                export_id,
                status="failed",
                elapsed_ms=int((perf_counter() - started_at) * 1000),
                error_type=type(exc).__name__,
                error_message=str(exc),
            )
            if isinstance(exc, QueryExportError):
                raise
            raise QueryExportError("Failed to generate the export file.", status_code=500) from exc

    def _build_export_sql(self, reviewed_sql: str) -> str:
        try:
            statements = [
                statement
                for statement in sqlglot.parse(reviewed_sql, read="postgres")
                if statement
            ]
        except sqlglot.errors.ParseError as exc:
            raise QueryExportError(
                "Reviewed SQL could not be parsed for export.", status_code=422
            ) from exc
        if len(statements) != 1:
            raise QueryExportError(
                "Export requires exactly one reviewed SQL statement.", status_code=422
            )
        expression = statements[0].copy()
        expression.set("limit", None)
        return expression.sql(dialect="postgres")

    def _serialize(
        self,
        source: QueryExportSource,
        result: SQLExecutionResult,
        export_format: ExportFormat,
    ) -> QueryExportArtifact:
        timestamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
        filename = f"finance-agent-{str(source.query_id)[:8]}-{timestamp}.{export_format}"
        if export_format == "xlsx":
            content = self._xlsx_bytes(source, result)
        elif export_format == "csv":
            content = self._csv_bytes(result)
        else:
            content = self._json_bytes(source, result)
        return QueryExportArtifact(
            content=content,
            media_type=self.media_types[export_format],
            filename=filename,
            row_count=result.row_count,
            truncated=result.truncated,
        )

    def _xlsx_bytes(self, source: QueryExportSource, result: SQLExecutionResult) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Query Result"
        sheet.freeze_panes = "A2"
        last_column = get_column_letter(max(len(result.columns), 1))
        last_row = max(result.row_count + 1, 1)
        sheet.auto_filter.ref = f"A1:{last_column}{last_row}"
        header_fill = PatternFill("solid", fgColor="E2E8F0")
        for column_index, column in enumerate(result.columns, start=1):
            cell = sheet.cell(row=1, column=column_index, value=column)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        widths = [len(str(column)) for column in result.columns]
        for row in result.rows:
            values = [self._safe_cell(row.get(column)) for column in result.columns]
            sheet.append(values)
            for index, value in enumerate(values):
                widths[index] = min(max(widths[index], len(str(value or ""))), 48)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = max(10, width + 2)

        metadata = workbook.create_sheet("Export Info")
        metadata.append(["Field", "Value"])
        metadata.append(["Query ID", str(source.query_id)])
        metadata.append(["Question", source.question])
        metadata.append(["Exported rows", result.row_count])
        metadata.append(["Truncated", "Yes" if result.truncated else "No"])
        metadata.append(["Export row limit", self.max_rows])
        metadata.column_dimensions["A"].width = 20
        metadata.column_dimensions["B"].width = 80
        for cell in metadata[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _csv_bytes(self, result: SQLExecutionResult) -> bytes:
        output = io.StringIO(newline="")
        writer = csv.DictWriter(output, fieldnames=result.columns, extrasaction="ignore")
        writer.writeheader()
        for row in result.rows:
            writer.writerow({column: self._safe_cell(row.get(column)) for column in result.columns})
        return ("\ufeff" + output.getvalue()).encode("utf-8")

    def _json_bytes(self, source: QueryExportSource, result: SQLExecutionResult) -> bytes:
        payload = {
            "query_id": str(source.query_id),
            "question": source.question,
            "row_count": result.row_count,
            "truncated": result.truncated,
            "export_row_limit": self.max_rows,
            "columns": result.columns,
            "rows": result.rows,
        }
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    @staticmethod
    def _safe_cell(value: Any) -> Any:
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value
